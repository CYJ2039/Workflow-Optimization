import cv2
import pytesseract
import re
import openpyxl
from datetime import datetime  # 👈 新增這行，用來處理時間格式

# ==========================================
# ⚙️ 系統設定區
# ==========================================
# 請確認 Tesseract 的安裝路徑是否正確
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# 檔案路徑設定
IMAGE_PATH = "test_card.png"      # 你的打卡紙圖片檔名
EXCEL_PATH = "count_salary.xlsx"  # 你的薪資計算 Excel 檔名
SHEET_NAME = "測試-加班費計算表"            # 要寫入的工作表名稱

# ==========================================
# 關卡一：影像預處理（去除雜訊，轉為黑白）
# ==========================================
def preprocess_image(image_path):
    img = cv2.imread(image_path)
    if img is None:
        print(f"❌ 錯誤：找不到圖片 {image_path}")
        return None
        
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    _, thresh = cv2.threshold(gray, 127, 255, cv2.THRESH_BINARY)
    return thresh

# ==========================================
# 關卡二與三：逐日切條掃描 ＆ 寫入 Excel
# ==========================================
def process_and_export(cleaned_img, excel_path):
    print("📂 正在開啟 Excel 檔案...")
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb[SHEET_NAME]
    except FileNotFoundError:
        print(f"❌ 錯誤：找不到 Excel 檔案 {excel_path}")
        return

    print("🔍 開始逐日掃描打卡紙並寫入資料...\n")
    
    # --- 📐 核心座標設定（需根據你的真實圖片微調） ---
    # 假設 1 號那列的起始 Y 座標是 150，每一天（列）的高度是 35 像素
    # 假設打卡紙的總寬度是 400 像素
    start_y = 453   
    row_height = 63
    img_width = 900 
    
    # # 我們設定迴圈跑 5 天（供測試用），之後可以改成 31 天
    # for day in range(1, 31):
    #     # 1. 算出每一天的 Y 軸範圍，把那一整條橫向切下來
    #     y_pos = start_y + (day - 1) * row_height
    #     day_row_img = cleaned_img[y_pos : y_pos + row_height, 0 : img_width]
        
    #     # 2. 呼叫 Tesseract 讀取這一條裡面的所有文字
    #     custom_config = r'--psm 6'
    #     text = pytesseract.image_to_string(day_row_img, config=custom_config)
        
    #     # 3. 啟動 Regex 智慧雷達，抓出格式為 HH:MM 的時間
    #     time_pattern = r'\d{2}:\d{2}'
    #     found_times = re.findall(time_pattern, text)
        
    #     # 4. 判斷打卡狀況，並指派給上班與下班
    #     start_time = ""
    #     end_time = ""
        
    #     if len(found_times) >= 2:
    #         start_time = found_times[0] # 抓到的第一個時間當上班
    #         end_time = found_times[-1]  # 抓到的最後一個時間當下班
    #         status = f"✅ 正常打卡 ({start_time} - {end_time})"
    #     elif len(found_times) == 1:
    #         start_time = found_times[0]
    #         status = f"⚠️ 漏打卡 (僅有 {start_time})"
    #     else:
    #         status = "❌ 無打卡紀錄"

    #     # 5. 精準寫入 Excel 欄位
    #     # Excel 的 4/1 在第 3 列，所以公式是：day + 2
    #     excel_row = day + 2
    #     ws[f"E{excel_row}"] = start_time  # E 欄是上班
    #     ws[f"D{excel_row}"] = end_time    # D 欄是下班
        
    #     print(f"📅 4/{day} (Excel 第 {excel_row} 列) -> {status}")

    # # 存檔並關閉
    # wb.save(excel_path)
    # print("\n🎉 全部處理完畢！請打開 Excel 看看自動連動的加班費吧！")

    # 🌟 升級 1：不要自己猜寬度，讓程式自動抓這張圖片的「真實寬度」
    img_height, img_width = cleaned_img.shape 
    
    for day in range(1, 16):
        y_pos = start_y + (day - 1) * row_height
        
        # 🌟 升級 2：切圖前先檢查，如果刀子已經超出圖片底部，就提早結束
        if y_pos >= img_height:
            print(f"⚠️ 警告：第 {day} 天的位置 ({y_pos}) 已經超出圖片總高度 ({img_height})，停止切圖！")
            break
            
        # 切下那一天的橫條
        day_row_img = cleaned_img[y_pos : y_pos + row_height, 0 : img_width]
        
        # 🌟 升級 3：二次確認切下來的圖片是不是空的
        if day_row_img.size == 0:
            continue
            
        # 呼叫 Tesseract 讀取這一條裡面的所有文字
        custom_config = r'--psm 6'
        text = pytesseract.image_to_string(day_row_img, config=custom_config)
        
        # --- (下面維持原本的 Regex 雷達抓時間邏輯) ---
        time_pattern = r'\d{2}:\d{2}'
        found_times = re.findall(time_pattern, text)
        
        # 判斷打卡狀況，並指派給上班與下班
        start_time = ""
        end_time = ""
        
        if len(found_times) >= 2:
            start_time = found_times[0] 
            end_time = found_times[-1]  
            status = f"✅ 正常打卡 ({start_time} - {end_time})"
        elif len(found_times) == 1:
            start_time = found_times[0]
            status = f"⚠️ 漏打卡 (僅有 {start_time})"
        else:
            status = "❌ 無打卡紀錄"

        # 5. 精準寫入 Excel 欄位
# ==========================================
        # 5. 精準寫入 Excel 欄位（防彈升級版）
        # ==========================================
        excel_row = day + 2
        
        # --- 處理上班時間 ---
        if start_time:
            try:
                # 嘗試轉換為標準時間
                ws[f"E{excel_row}"] = datetime.strptime(start_time, "%H:%M").time()
            except ValueError:
                # 如果辨識出 99:02 這種火星時間，抓到這裡防崩潰
                print(f"⚠️  【辨識異常】4/{day} 上班時間抓到怪怪的字 '{start_time}'，已先改以文字寫入。")
                ws[f"E{excel_row}"] = start_time
        else:
            ws[f"E{excel_row}"] = ""
            
        # --- 處理下班時間 ---
        if end_time:
            try:
                # 嘗試轉換為標準時間
                ws[f"D{excel_row}"] = datetime.strptime(end_time, "%H:%M").time()
            except ValueError:
                print(f"⚠️  【辨識異常】4/{day} 下班時間抓到怪怪的字 '{end_time}'，已先改以文字寫入。")
                ws[f"D{excel_row}"] = end_time
        else:
            ws[f"D{excel_row}"] = ""
        
        print(f"📅 4/{day} (Excel 第 {excel_row} 列) -> {status}")

    # 存檔並關閉
    wb.save(excel_path)
    print("\n🎉 全部處理完畢！請打開 Excel 看看自動連動的加班費吧！")

# ==========================================
# 🚀 總指揮官（主程式）
# ==========================================
if __name__ == "__main__":
    processed_img = preprocess_image(IMAGE_PATH)
    if processed_img is not None:
        process_and_export(processed_img, EXCEL_PATH)